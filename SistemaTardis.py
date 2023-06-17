import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl
import pathlib
from PIL import Image, ImageTk


janela = ctk.CTk()

class AppSOD():
    def __init__(self):
        self.janela=janela
        self.tema()
        self.tela()
        self.criação_xlsx()
        janela.mainloop()

    def tema(self):    
        ctk.set_appearance_mode('dark')
        ctk.set_default_color_theme('dark-blue')

    def tela(self):
        janela.geometry('915x550')
        janela.title('Matriz SOD')
        janela.iconbitmap('matrix2.ico')
        janela.resizable(False, False)
    

    def criação_xlsx(self):  
        arquivo_xlsx = pathlib.Path('MatrizSOD.xlsx')
        if arquivo_xlsx.exists():
            pass
        else:
            arquivo_xlsx=openpyxl.Workbook()
            arquivo_xlsx.create_sheet('Matriz', 0)
            Matriz=arquivo_xlsx.active
            Matriz['A1']='X'

            arquivo_xlsx.create_sheet('Sistemas', 0)
            Sistemas=arquivo_xlsx.active
            Sistemas['A1']='Código do sistema'
            Sistemas['B1']='Nome do sistema'
            

            arquivo_xlsx.create_sheet('Perfis', 0)
            Perfis=arquivo_xlsx.active
            Perfis['A1']='Código do Sistema'
            Perfis['B1']='Nome do Perfil'
            Perfis['C1']='Descrição detalhada do perfil'
            Perfis['D1']='Código Sistema + Nome Perfil'


            arquivo_xlsx.create_sheet('Usuários', 0)
            Usuários=arquivo_xlsx.active
            Usuários['A1']='CPF'
            Usuários['B1']='Código do sistema'
            Usuários['C1']='Nome do Perfil'

            arquivo_xlsx.create_sheet('Motivos', 0)
            Motivos=arquivo_xlsx.active
            Motivos['A1']='Perfil 1'
            Motivos['B1']='Perfil 2'
            Motivos['C1']='Motivo do conflito'

            arquivo_xlsx.save('MatrizSOD.xlsx') 

        #Variáveis Globais e outros    
        siscoluna_b = []
        perfcoluna_d = []
        usucoluna_a = []
        arquivo_xlsx=openpyxl.load_workbook('MatrizSOD.xlsx')
        Sistemas=arquivo_xlsx['Sistemas']
        Perfis=arquivo_xlsx['Perfis']
        Matriz=arquivo_xlsx['Matriz']
        Usuários=arquivo_xlsx['Usuários']
        Motivos=arquivo_xlsx['Motivos']

        frameprin = ctk.CTkFrame(master=janela, width =915, height= 475, bg_color='#1f2a3e',corner_radius=30, border_width= 5, border_color='#1f2a3e')
        frameprin.place(x=0, y=55)
        


        def fechar_frames():
            for widget in frameprin.winfo_children():
                if isinstance(widget, ctk.CTkFrame):
                    widget.destroy()

         

        def criar_listas():         
            # Combobox - Código de Sistemas
            for cellsis in Sistemas['B']:
                if cellsis.value is not None:
                    siscoluna_b.append(cellsis.value)
            siscoluna_b[0] = "Escolha o Sistema"

            # Combobox - Código de Perfil
            for cellper in Perfis['D']:
                if cellper.value is not None:
                    perfcoluna_d.append(cellper.value)
            perfcoluna_d[0] = "Escolha o Perfil"

            # Combobox - Usuário
            for cellusu in Usuários['A']:
                if cellusu.value is not None:
                    usucoluna_a.append(cellusu.value)
            usucoluna_a[0] = "Escolha o Usuário"

    
        def frames_inicio():
            fechar_frames()

            #Cabeçalho
            cabeçalho_frame = ctk.CTkFrame(master= janela, width=915, height= 55, bg_color="#1f2a3e", fg_color="#1f2a3e")
            cabeçalho_frame.place(x=0,y=0)
            titulo_label = ctk.CTkLabel(master= cabeçalho_frame, text= "Sistema de Segregação de Funções", font=('Century Ghotic bold', 38))
            titulo_label.place(x=150, y=10)

            #Página Inicial
            framecadastros = ctk.CTkFrame(master=frameprin, width =915, height= 475, bg_color='#1f2a3e',corner_radius=30, border_width= 5, border_color='#1f2a3e', fg_color='#212121')
            framecadastros.place(x=0, y=0)
            image = Image.open("tardis_image.png")
            image = image.resize((300, 300), Image.LANCZOS)
            photo = ImageTk.PhotoImage(image)

            image_label = Label(framecadastros, image=photo, bg="#212121")
            image_label.image = photo
            image_label.place(x=307.5, y=87.5)
            #lb_cadastros=ctk.CTkLabel(master=framecadastros, text="Escolha uma opção abaixo\npara fazer um cadastro.", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=85, y=15)
            btn_sistemacad=ctk.CTkButton(master=framecadastros, width= 200,text='Cadastrar Sistema'.upper(),command=cad_sis, fg_color='#245983', hover_color='#063970').place(x=68.75, y=104.75)
            btn_perfilcad=ctk.CTkButton(master=framecadastros, width= 200,text='Cadastrar Perfil de Acesso'.upper(),command=cad_per, fg_color='#245983', hover_color='#063970').place(x=646.25, y=104.75)
            btn_matrizcad=ctk.CTkButton(master=framecadastros, width= 200 ,text='Cadastrar Conflitos'.upper(),command=cad_mat, fg_color='#245983', hover_color='#063970').place(x=38.75, y=223.5)
            btn_usuariocad=ctk.CTkButton(master=framecadastros, width= 200 ,text='Cadastrar Usuário'.upper(),command=cad_usu, fg_color='#245983', hover_color='#063970').place(x=676.25, y=223.5)
            btn_matrizcon=ctk.CTkButton(master=framecadastros, width= 200,text='Consultar Conflitos/Perfis'.upper(),command=con_mat, fg_color='#245983', hover_color='#063970').place(x=68.75, y=342.25)
            btn_usuariocon=ctk.CTkButton(master=framecadastros, width= 200, text='Consultar Usuário'.upper(),command=con_usu, fg_color='#245983', hover_color='#063970').place(x=646.25, y=342.25)
            
            #Rodapé
            rodape_frame = ctk.CTkFrame(master= janela, width=915, height= 25, bg_color="#1f2a3e", fg_color="#1f2a3e")
            rodape_frame.place(x= 0, y=525)
            dev_label = ctk.CTkLabel(master= rodape_frame, text= "Developed by: TARdis DEVs™", font=('Century Ghotic bold',11))
            dev_label.place(x=5 , y=0)
            version_label = ctk.CTkLabel(master= rodape_frame, text= "V1.0.1", font=('Century Ghotic bold', 11))
            version_label.place(x= 875, y= 0)

        # Cadastro dos sistemas (código do sistema e nome do sistema)
        def cad_sis():
            fechar_frames()

            frame1=ctk.CTkFrame(master=frameprin, width =915, height= 475, bg_color='#1f2a3e',corner_radius=30, border_width= 5, border_color='#1f2a3e', fg_color='#212121')
            frame1.pack(side=TOP)
            frame_faixa=ctk.CTkFrame(master=frame1, width=640,height=50,corner_radius=12,bg_color='#212121',fg_color='#245983')
            frame_faixa.place(x=137.5,y=17)         
            btn_voltar=ctk.CTkButton(master=frame1, text='Voltar'.upper(),command=frames_inicio, fg_color='#245983', hover_color='#063970').place(x=640, y=430)       
            
            title=ctk.CTkLabel(frame_faixa, text="Cadastro dos Sistemas",font=('Century Ghotic bold', 24),text_color="#fff").place(x=230,y=10)
            #span=ctk.CTkLabel(master=frame1, text="Preencha todos os campos do formulário!", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=195, y=60)

            def lista_sis():
                listasis = ctk.CTkTextbox(frame1)
                listasis.place(x=517.5, y=130)

                for siscod in Sistemas.iter_rows(min_row=2, values_only=True):
                    sistexto = str(siscod[0]) + ' - ' + str(siscod[1]) + '\n'
                    listasis.insert(END, sistexto)      
                listasis.configure(state=DISABLED)
            
            lista_sis()
            
            def submit():
                cod_sis = codsis_value.get()
                nome_sis = nomesis_value.get()
                temounaotem=0
                temounaotem2=0
                for cell in Sistemas['A']:
                    if cell.value == cod_sis:
                        temounaotem = 1
                for cell in Sistemas['B']:
                    if cell.value == nome_sis:
                        temounaotem2 = 1
                        

                if (cod_sis=='' or nome_sis==''):
                    messagebox.showerror('Sistema', 'ERRO\nPor favor preencha todos os campos!')

                elif (temounaotem==1 and temounaotem2==1):
                    messagebox.showerror('Sistema', 'CÓDIGO E NOME EXISTENTE\nJá existe esse CÓDIGO e esse NOME cadastrado!')

                elif (temounaotem == 1):
                    messagebox.showerror('Sistema', 'CÓDIGO EXISTENTE\nJá existe esse CÓDIGO cadastrado!')

                elif (temounaotem2 == 1):
                    messagebox.showerror('Sistema', 'NOME EXISTENTE\nJá existe esse NOME cadastrado!')

                else:
                    Sistemas.cell(column=1, row=Sistemas.max_row+1, value=cod_sis)
                    Sistemas.cell(column=2, row=Sistemas.max_row, value=nome_sis)
                    arquivo_xlsx.save(r'MatrizSOD.xlsx')
                    messagebox.showinfo('Sistema', 'Dados salvos com sucesso!')
                    siscoluna_b.append(nome_sis)
                    clear()
                    lista_sis()

            

            def clear():
                codsis_value.set('')
                nomesis_value.set('')
                    

            codsis_value=StringVar()
            nomesis_value=StringVar()
            codsis_entry=ctk.CTkEntry(master=frame1, width=200, textvariable=codsis_value, font=('Century Gohtic', 16),fg_color="transparent").place(x=187.5, y=150)
            nomesis_entry=ctk.CTkEntry(master=frame1, width=200, textvariable=nomesis_value, font=('Century Gohtic', 16),fg_color="transparent").place(x=187.5, y=230)
            lb_codsis=ctk.CTkLabel(master=frame1, text="Código do Sistema:", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=187.5, y=120)
            lb_nomesis=ctk.CTkLabel(master=frame1, text="Nome do Sistema:", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=187.5, y=200)
            btn_submit=ctk.CTkButton(master=frame1, text='Salvar dados'.upper(),command=submit, fg_color='#245983', hover_color='#063970').place(x=137.5, y=430)
            btn_clear=ctk.CTkButton(master=frame1, text='Limpar dados'.upper(),command=clear, fg_color='#245983', hover_color='#063970').place(x=297.5, y=430)
            lb_listasis=ctk.CTkLabel(master=frame1, text='Lista de Sistemas:', font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=517.5, y=100)
            
            

        # Cadastros dos Perfis (Código do Sistema, Nome do Perfil e Descrição detalhada do perfil)
        def cad_per():
            fechar_frames()
            frame2=ctk.CTkFrame(master=frameprin, width =915, height= 475, bg_color='#1f2a3e',corner_radius=30, border_width= 5, border_color='#1f2a3e', fg_color='#212121')
            frame2.pack(side=TOP)
            btn_voltar=ctk.CTkButton(master=frame2, text='Voltar'.upper(),command=frames_inicio, fg_color='#245983', hover_color='#063970').place(x=640, y=430)       
            frame_faixa=ctk.CTkFrame(master=frame2, width=640,height=50,corner_radius=12,bg_color='#212121',fg_color='#245983')
            frame_faixa.place(x=137.5,y=17) 
            title=ctk.CTkLabel(frame_faixa, text="Cadastro dos Perfis",font=('Century Ghotic bold', 24),text_color="#fff").place(x=230,y=10)
            #span=ctk.CTkLabel(master=frame2, text="Preencha todos os campos do formulário!", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=195, y=60)
            
            combosis = ctk.CTkComboBox(frame2, values=siscoluna_b, width=220, state="readonly")
            combosis.place(x=177.5, y=130)
            combosis.set(siscoluna_b[0])

            def submit():
                codsis = combosis.get()
                nomeperf = nomeperf_value.get()
                cod_perf = codsis + ' - ' + nomeperf
                descrição = desc_entry.get(0.0, END)
                temesseperfil=0
                for cell in Perfis['D']:
                    if cell.value == cod_perf:
                        temesseperfil = 1
                        break
                        

                if (codsis=="Escolha o Sistema" and nomeperf==''):
                    messagebox.showerror('Sistema', 'ERRO\nPor favor preencha os campos:\nSistema e Nome do Perfil!')

                elif (codsis=="Escolha o Sistema"):
                    messagebox.showerror('Sistema', 'ERRO\nPor favor escolha o Sistema!')

                elif (nomeperf==''):
                    messagebox.showerror('Sistema', 'ERRO\nPor favor preencha o campo: Nome do Perfil!')

                elif (temesseperfil == 1):
                    messagebox.showerror('Sistema', 'PERFIL EXISTENTE\nJá existe esse PERFIL cadastrado!')
                   
                else:
                    Perfis.cell(column=1, row=Perfis.max_row+1, value=codsis)
                    Perfis.cell(column=2, row=Perfis.max_row, value=nomeperf)
                    Perfis.cell(column=3, row=Perfis.max_row, value=descrição)
                    Perfis.cell(column=4, row=Perfis.max_row, value=cod_perf)
                    Matriz.cell(column=1, row=Matriz.max_row+1, value=cod_perf)
                    Matriz.cell(column=Matriz.max_column+1, row=1, value=cod_perf)
                    
                    arquivo_xlsx.save(r'MatrizSOD.xlsx')
                    messagebox.showinfo('Sistema', 'Dados salvos com sucesso!')
                    perfcoluna_d.append(cod_perf)
                    clear()
                    lista_perf()
            

            def clear():
                combosis.set(siscoluna_b[0])
                nomeperf_value.set('')
                desc_entry.delete(0.0, END)
                    

            nomeperf_value=StringVar()
            nomeperf_entry=ctk.CTkEntry(master=frame2, width=220, textvariable=nomeperf_value, font=('Century Gohtic', 16),fg_color="transparent").place(x=177.5, y=210)
            desc_entry=ctk.CTkTextbox(master=frame2, wrap=WORD ,width=220, height=110, font=("arial", 14), border_color="#aaa", border_width=2, fg_color="transparent")
            desc_entry.place(x=177.5, y=290)
            lb_desc=ctk.CTkLabel(master=frame2, text="Descrição detalhada do perfil:", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=177.5, y=260)
            lb_perf=ctk.CTkLabel(master=frame2, text="Sistema:", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=177.5, y=100)
            lb_nomeperf=ctk.CTkLabel(master=frame2, text="Nome do Perfil:", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=177.5, y=180)
            btn_submit=ctk.CTkButton(master=frame2, text='Salvar dados'.upper(),command=submit, fg_color='#245983', hover_color='#063970').place(x=137.5, y=430)
            btn_clear=ctk.CTkButton(master=frame2, text='Limpar dados'.upper(),command=clear, fg_color='#245983', hover_color='#063970').place(x=297.5, y=430)
            lb_listasis=ctk.CTkLabel(master=frame2, text='Lista de Perfis:', font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=517.5, y=100)

            def lista_perf():
                listaperf = ctk.CTkTextbox(frame2, height= 272.5, width=230)
                listaperf.place(x=517.5, y=130)

                for perf in Perfis.iter_rows(min_row=2, values_only=True):
                    perftexto = str(perf[3]) + '\n'
                    listaperf.insert(END, perftexto)
                    
                listaperf.configure(state=DISABLED)
            
            lista_perf()


        # Cadastro da Matriz SOD (Código do Sistema 1, Nome do Perfil 1, Código do Sistema 2 e Nome do Perfil 2)
        def cad_mat():
            fechar_frames()
            frame3=ctk.CTkFrame(master=frameprin, width =915, height= 475, bg_color='#1f2a3e',corner_radius=30, border_width= 5, border_color='#1f2a3e', fg_color='#212121')
            frame3.pack(side=TOP)            
            btn_voltar=ctk.CTkButton(master=frame3, text='Voltar'.upper(),command=frames_inicio, fg_color='#245983', hover_color='#063970').place(x=640, y=430)        
            frame_faixa=ctk.CTkFrame(master=frame3, width=640,height=50,corner_radius=12,bg_color='#212121',fg_color='#245983')
            frame_faixa.place(x=137.5,y=17) 
            title=ctk.CTkLabel(frame_faixa, text="Cadastro da Matriz SOD",font=('Century Ghotic bold', 24),text_color="#fff").place(x=188,y=10)
            span=ctk.CTkLabel(master=frame3, text="Escolha dois Perfis que terão conflitos entre si!", font=("Century Gothic bold", 18), text_color=["#000","#fff"]).place(x=271, y=80)
            


            def submit():
                perf = comboperf.get()
                perf2 = comboperf2.get()
                motivo = why_entry.get(0.0, END)
                    
                if (perf=='Escolha o Perfil' and perf2=='Escolha o Perfil'):
                    messagebox.showerror('Sistema', 'ERRO\nPor favor escolha os perfis!')

                elif (perf!='Escolha o Perfil' and perf2=='Escolha o Perfil'):
                    messagebox.showerror('Sistema', 'ERRO\nPor favor escolha o Perfil 2!')

                elif (perf=='Escolha o Perfil' and perf2!='Escolha o Perfil'):
                    messagebox.showerror('Sistema', 'ERRO\nPor favor escolha o Perfil 1!')

                elif (perf == perf2):
                    messagebox.showerror('Sistema', 'ERRO\nNão pode ter conflito entre o Perfil e ele mesmo!')

                else:
                    linha_celula = None
                    coluna_celula = None 
                    for linha in range(1, Matriz.max_row+1):
                        valor_celula = Matriz.cell(row=linha, column=1).value
                        if valor_celula == perf:
                            linha_celula = linha
                            break
                    for coluna in range(1, Matriz.max_column+1):
                        valor_celula = Matriz.cell(row=1, column=coluna).value
                        if valor_celula == perf2:
                            coluna_celula = coluna
                            break
                    
                    if Matriz.cell(column=coluna_celula, row=linha_celula).value or Matriz.cell(column=linha_celula, row=coluna_celula) == 1:
                        messagebox.showerror('Sistema', 'ERRO\nEsse conflito já foi cadastrado!')

                    elif linha_celula is not None and coluna_celula is not None:
                        Matriz.cell(column=coluna_celula, row=linha_celula, value=1)
                        Matriz.cell(column=linha_celula, row=coluna_celula, value=1)
                        Motivos.cell(column=1, row=Motivos.max_row+1, value=perf)
                        Motivos.cell(column=2, row=Motivos.max_row, value=perf2)
                        Motivos.cell(column=3, row=Motivos.max_row, value=motivo)
                        Motivos.cell(column=1, row=Motivos.max_row+1, value=perf2)
                        Motivos.cell(column=2, row=Motivos.max_row, value=perf)
                        Motivos.cell(column=3, row=Motivos.max_row, value=motivo)
                        arquivo_xlsx.save(r'MatrizSOD.xlsx')
                        messagebox.showinfo('Sistema', 'Dados salvos com sucesso!')
                        clear()    
                

            def clear():
                comboperf.set(perfcoluna_d[0])
                comboperf2.set(perfcoluna_d[0])
                why_entry.delete(0.0, END)
                    
            comboperf = ctk.CTkComboBox(frame3, values=perfcoluna_d, width=300, state="readonly")
            comboperf.place(x=307.5, y=150)
            comboperf.set(perfcoluna_d[0])
            comboperf2 = ctk.CTkComboBox(frame3, values=perfcoluna_d, width=300, state="readonly")
            comboperf2.place(x=307.5, y=230)
            comboperf2.set(perfcoluna_d[0])

            perf=StringVar()
            perf2=StringVar()
            why_entry=ctk.CTkTextbox(master=frame3, wrap=WORD ,width=300, height=90, font=("arial", 14), border_color="#aaa", border_width=2, fg_color="transparent")
            why_entry.place(x=307.5, y=310)
            lb_desc=ctk.CTkLabel(master=frame3, text="Motivo do conflito:", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=307.5, y=280)
            lb_perf=ctk.CTkLabel(master=frame3, text="Nome do Perfil 1:", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=307.5, y=120)
            lb_perf2=ctk.CTkLabel(master=frame3, text="Nome do Perfil 2:", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=307.5, y=200)
            btn_submit=ctk.CTkButton(master=frame3, text='Salvar dados'.upper(),command=submit, fg_color='#245983', hover_color='#063970').place(x=137.5, y=430)
            btn_clear=ctk.CTkButton(master=frame3, text='Limpar dados'.upper(),command=clear, fg_color='#245983', hover_color='#063970').place(x=387.5, y=430)
  


        # Cadastro dos Usuários (CPF do Usuário, Código do Sistema e Nome do perfil)
        def cad_usu():
            fechar_frames()
            fram4=ctk.CTkFrame(master=frameprin, width =915, height= 475, bg_color='#1f2a3e',corner_radius=30, border_width= 5, border_color='#1f2a3e', fg_color='#212121')
            fram4.pack(side=TOP)            
            btn_voltar=ctk.CTkButton(master=fram4, text='Voltar'.upper(),command=frames_inicio, fg_color='#245983', hover_color='#063970').place(x=640, y=430)     
            frame_faixa=ctk.CTkFrame(master=fram4, width=640,height=50,corner_radius=12,bg_color='#212121',fg_color='#245983')
            frame_faixa.place(x=137.5,y=17) 
            title=ctk.CTkLabel(frame_faixa, text="Cadastro dos Usuários",font=('Century Ghotic bold', 24),text_color="#fff").place(x=200,y=10)
            span=ctk.CTkLabel(master=fram4, text="Preencha todos os campos do formulário!", font=("Century Gothic bold", 18), text_color=["#000","#fff"]).place(x=290, y=80)

            comboperf = ctk.CTkComboBox(fram4, values=perfcoluna_d, width=300, state="readonly")
            comboperf.place(x=307.5, y=290)
            comboperf.set(perfcoluna_d[0])

            def submit():
                cpf = cpf_value.get()
                perf = comboperf.get()

                temcpf = temperf = 0
                for cell in Usuários['A']:
                        if cell.value == cpf:
                            temcpf = 1
                            break
                 

                def cadastrar_usu():
                    nome = nome_value.get()
                    for linha34 in range(1, Usuários.max_row+2):
                        valor_celula34 = Usuários.cell(row=linha34, column=3).value
                        if valor_celula34 == None:
                            Usuários.cell(column=1, row=linha34, value=cpf)
                            Usuários.cell(column=2, row=linha34, value=nome)
                            Usuários.cell(column=3, row=linha34, value=perf)
                            if temcpf == 0:
                                Usuários.cell(column=4, row=linha34, value='1ªx')                                     
                            arquivo_xlsx.save(r'MatrizSOD.xlsx')
                            messagebox.showinfo('Sistema', 'Dados salvos com sucesso!')
                            break
                    cad_usu()
                                           
                if (cpf=='' or perf=='Escolha o Perfil'):
                    messagebox.showerror('Sistema', 'ERRO\nPor favor preencha todos os campos!')
                
                elif temcpf == 1:
                    for linha1 in range(1, Usuários.max_row+2):
                        valor_celula1 = Usuários.cell(row=linha1, column=3).value
                        if valor_celula1 == perf and Usuários.cell(row=linha1, column=1).value == cpf:
                            temperf = 1
                            break
                    if temperf == 1:
                        messagebox.showerror('Sistema', 'ERRO\nEste Perfil já foi cadastrado a esse Usuário!')
                    
                    else:
                        for linha in range(1, Usuários.max_row+2):
                            valor_celula = Usuários.cell(row=linha, column=1).value
                            if valor_celula == cpf:
                                perf2 = Usuários.cell(row=linha, column=3).value
                                break
                        linha_celula = None
                        coluna_celula = None 
                        for linha in range(1, Matriz.max_row+1):
                            valor_celula = Matriz.cell(row=linha, column=1).value
                            if valor_celula == perf:
                                linha_celula = linha
                                break
                        for coluna in range(1, Matriz.max_column+1):
                            valor_celula = Matriz.cell(row=1, column=coluna).value
                            if valor_celula == perf2:
                                coluna_celula = coluna
                                break
                        
                        motivotexto = StringVar
                        if Matriz.cell(column=coluna_celula, row=linha_celula).value or Matriz.cell(column=linha_celula, row=coluna_celula) == 1:
                            valor1 = Matriz.cell(row=1, column=coluna_celula).value
                            valor2 = Matriz.cell(row=linha_celula, column=1).value
                            for perf1 in Motivos['A']:
                                veriperf = perf1.value
                                if veriperf == valor1 and valor2 == Motivos.cell(row=perf1.row, column=2).value:
                                    motivotexto = Motivos.cell(row=perf1.row, column=3).value
                                    messagebox.showerror('Sistema', 'ERRO\nEste Perfil entra em conflito com outro Perfil já cadastrado neste Usuário!\nMOTIVO: ' + motivotexto)
                                    break
                        
                        else:
                            cadastrar_usu()       

                else:
                    framenome=ctk.CTkFrame(master=fram4, width =915, height= 475, bg_color='#1f2a3e',corner_radius=30, border_width= 5, border_color='#1f2a3e', fg_color='#212121')
                    framenome.place(x=0, y=0)            
                    btn_voltar=ctk.CTkButton(master=framenome, text='Voltar'.upper(),command=frames_inicio, fg_color='#245983', hover_color='#063970').place(x=640, y=430)     
                    frame_faixa=ctk.CTkFrame(master=framenome, width=640,height=50,corner_radius=12,bg_color='#212121',fg_color='#245983')
                    frame_faixa.place(x=137.5,y=17) 
                    title=ctk.CTkLabel(frame_faixa, text="Cadastro dos Usuários",font=('Century Ghotic bold', 24),text_color="#fff").place(x=200,y=10)
                    span1=ctk.CTkLabel(master=framenome, text="NOVO USUÁRIO!!!", font=("Century Gothic bold", 22), text_color=["#000","#fff"]).place(x=290, y=80)
                    span=ctk.CTkLabel(master=framenome, text="Preencha com o nome completo do Usuário!", font=("Century Gothic bold", 18), text_color=["#000","#fff"]).place(x=290, y=110)
                    nome_entry = ctk.CTkEntry(master=framenome, width=300, textvariable=nome_value, font=('Century Gohtic', 16),fg_color="transparent").place(x=307.5, y=230)
                    lb_name=ctk.CTkLabel(master=framenome, text="Nome do Usuário:", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=307.5, y=200)
                    btn_submit=ctk.CTkButton(master=framenome, text='Salvar dados'.upper(),command=cadastrar_usu, fg_color='#245983', hover_color='#063970').place(x=137.5, y=430)  
                    btn_voltar=ctk.CTkButton(master=fram4, text='Voltar'.upper(),command=cad_usu, fg_color='#245983', hover_color='#063970').place(x=640, y=430)

            
                

            def clear():
                cpf_value.set('')
                nome_value.set('')
                comboperf.set(perfcoluna_d[0])
                
                                    
            cpf_value=StringVar()
            nome_value=StringVar()
            cpf_entry=ctk.CTkEntry(master=fram4, width=200, textvariable=cpf_value, font=('Century Gohtic', 16),fg_color="transparent").place(x=357.5, y=170)
            lb_cpf=ctk.CTkLabel(master=fram4, text="CPF do Usuário:", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=357.5, y=140)
            lb_perf=ctk.CTkLabel(master=fram4, text="Nome do Perfil:", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=307.5, y=260)
            btn_submit=ctk.CTkButton(master=fram4, text='Salvar dados'.upper(),command=submit, fg_color='#245983', hover_color='#063970').place(x=137.5, y=430)
            btn_clear=ctk.CTkButton(master=fram4, text='Limpar dados'.upper(),command=clear, fg_color='#245983', hover_color='#063970').place(x=387.5, y=430)


        def con_mat():
            fechar_frames()
            frame34=ctk.CTkFrame(master=frameprin, width =915, height= 475, bg_color='#1f2a3e',corner_radius=30, border_width= 5, border_color='#1f2a3e', fg_color='#212121')
            frame34.pack(side=TOP)            
            btn_voltar=ctk.CTkButton(master=frame34, text='Voltar'.upper(),command=frames_inicio, fg_color='#245983', hover_color='#063970').place(x=640, y=430)    
            frame_faixa=ctk.CTkFrame(master=frame34, width=640,height=50,corner_radius=12,bg_color='#212121',fg_color='#245983')
            frame_faixa.place(x=137.5,y=17) 
            title=ctk.CTkLabel(frame_faixa, text="Consultar os Conflitos de cada Perfil",font=('Century Ghotic bold', 24),text_color="#fff").place(x=120,y=10)
            lb_pesquise=ctk.CTkLabel(master=frame34, text="Escolha abaixo o Perfil\nque deseja consultar:", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=195, y=140)
            lb_listasis=ctk.CTkLabel(master=frame34, text='Lista de Perfis:', font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=517.5, y=100)

            listasis = ctk.CTkTextbox(frame34, height=272.5)
            listasis.place(x=517.5, y=130)

            for perf in Perfis.iter_rows(min_row=2, values_only=True):
                perftexto = str(perf[3]) + '\n'
                listasis.insert(END, perftexto)
                
            listasis.configure(state=DISABLED)

            def consultar_sis():
                def voltar_consis():
                    arquivo_xlsx=openpyxl.load_workbook('MatrizSOD.xlsx')
                    con_mat()
                codsis=comboperf.get()
                if codsis == "Escolha o Perfil":
                    messagebox.showerror('Sistema', 'Escolha um Perfil!')
                else:
                    fechar_frames()
                    frame_siscon=ctk.CTkFrame(master=frameprin, width =915, height= 475, bg_color='#1f2a3e',corner_radius=30, border_width= 5, border_color='#1f2a3e', fg_color='#212121')
                    frame_siscon.pack(side=TOP)
                    frame_faixa=ctk.CTkFrame(master=frame_siscon, width=640,height=50,corner_radius=12,bg_color='#212121',fg_color='#245983')
                    frame_faixa.place(x=137.5,y=17) 
                    title=ctk.CTkLabel(frame_faixa, text="Lista de Perfis Conflitantes",font=('Century Ghotic bold', 24),text_color="#fff").place(x=180,y=10)  
                    btn_vota=ctk.CTkButton(master=frame_siscon, text='Voltar'.upper(),command=voltar_consis, fg_color='#245983', hover_color='#063970')
                    btn_vota.place(x=640, y=430) 
                    lbl_titulo = ctk.CTkLabel(master=frame_siscon, text='Perfis conflitantes:', font=('Century Gothic bold', 14, 'bold'))
                    lbl_titulo.place(x=550.5, y=100)

                    lbl_titulo = ctk.CTkLabel(master=frame_siscon, text='Perfil selecionado:', font=('Century Gothic bold', 24, 'bold'))
                    lbl_titulo.place(x=190, y=120)
                    lbl_titulo = ctk.CTkLabel(master=frame_siscon, text=codsis, font=('Century Gothic bold', 22, 'bold'))
                    lbl_titulo.place(x=190, y=150)

                    lbl_titulo = ctk.CTkLabel(master=frame_siscon, text='Os perfis ao lado geram conflito\ncom o Perfil escolhido.', font=('Century Gothic bold', 14))
                    lbl_titulo.place(x=210, y=360)
                    

                    listaconflito = ctk.CTkTextbox(frame_siscon, height=272.5)
                    listaconflito.place(x=550.5, y=130)
                    

                    desc = ctk.CTkTextbox(master=frame_siscon, wrap=WORD, height=150, width= 250)
                    desc.place(x=190, y=190)

                    for cell in Perfis['D']:
                        if cell.value == codsis:
                            desctexto = str(Perfis.cell(row=cell.row, column=3).value)
                            desc.insert(END, desctexto)
                            break
                    
                    for conflito in range(1, Matriz.max_row+1):
                        valor_celula = Matriz.cell(row=conflito, column=1).value
                        if valor_celula == codsis:
                            for conflito1 in range(1, Matriz.max_column+1):
                                valor_celula1 = Matriz.cell(row=conflito, column=conflito1).value
                                if valor_celula1 == 1:
                                    conftexto = str(Matriz.cell(row=1, column=conflito1).value) + '\n'
                                    listaconflito.insert(END, conftexto)

                    listaconflito.configure(state=DISABLED)
                    desc.configure(state=DISABLED)

            comboperf = ctk.CTkComboBox(frame34, values=perfcoluna_d, width=200, state="readonly")
            comboperf.place(x=177.5, y= 190)
            comboperf.set(perfcoluna_d[0])
            btn_consultar=ctk.CTkButton(master=frame34, text='Consultar'.upper(),command=consultar_sis, fg_color='#245983', hover_color='#063970').place(x=206.5, y=230)

        def con_usu():
            fechar_frames()
            frame6=ctk.CTkFrame(master=frameprin, width =915, height= 475, bg_color='#1f2a3e',corner_radius=30, border_width= 5, border_color='#1f2a3e', fg_color='#212121')
            frame6.pack(side=TOP)            
            btn_voltar=ctk.CTkButton(master=frame6, text='Voltar'.upper(),command=frames_inicio, fg_color='#245983', hover_color='#063970').place(x=640, y=430)  
            frame_faixa=ctk.CTkFrame(master=frame6, width=640,height=50,corner_radius=12,bg_color='#212121',fg_color='#245983')
            frame_faixa.place(x=137.5,y=17)
            title=ctk.CTkLabel(frame_faixa, text="Consulta dos Usuários",font=('Century Ghotic bold', 24),text_color="#fff").place(x=170,y=10)
            lb_pesquise=ctk.CTkLabel(master=frame6, text="Escreva o CPF de um\nusuário já cadastrado:", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=195, y=140)
            lb_listasis=ctk.CTkLabel(master=frame6, text='Lista de Usuários:', font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=467.5, y=100)

            listausu = ctk.CTkTextbox(frame6, width= 300, height=272.5)
            listausu.place(x=467.5, y=130)

            for usu in Usuários.iter_rows(min_row=2, values_only=True):
                if usu[3] == '1ªx':
                    usutexto = str(usu[1]) + ' / CPF:  ' + str(usu[0]) + '\n'
                    listausu.insert(END, usutexto)
                
            listausu.configure(state=DISABLED)


            def consultar_usu():
                def voltar_consis():
                    arquivo_xlsx=openpyxl.load_workbook('MatrizSOD.xlsx')
                    con_usu()
                cpf=cpf_value2.get()
                temounaotem=0
                for cell in Usuários['A']:
                    if cell.value == cpf:
                        temounaotem = 1
                        break
                if temounaotem == 0:
                    messagebox.showerror('Sistema', 'Escreva o CPF de um usuário já cadastrado!')
                else:
                    fechar_frames()
                    frame_siscon=ctk.CTkFrame(master=frameprin, width =915, height= 475, bg_color='#1f2a3e',corner_radius=30, border_width= 5, border_color='#1f2a3e', fg_color='#212121')
                    frame_siscon.pack(side=TOP)
                    frame_faixa=ctk.CTkFrame(master=frame_siscon, width=640,height=50,corner_radius=12,bg_color='#212121',fg_color='#245983')
                    frame_faixa.place(x=137.5,y=17) 
                    title=ctk.CTkLabel(frame_faixa, text="Lista de Perfis do Usuário consultado",font=('Century Ghotic bold', 24),text_color="#fff").place(x=160,y=10)  
                    btn_vota=ctk.CTkButton(master=frame_siscon, text='Voltar'.upper(),command=voltar_consis, fg_color='#245983', hover_color='#063970')
                    btn_vota.place(x=640, y=430) 
                    lbl_titulo = ctk.CTkLabel(master=frame_siscon, text='Perfis do Usuário:', font=('Century Gothic bold', 14, 'bold'))
                    lbl_titulo.place(x=517.5, y=100)

                    listaperfis = ctk.CTkTextbox(frame_siscon, height=272.5, width= 240)
                    listaperfis.place(x=517.5, y=130)

                    japegounome = 0

                    for usu in range(1, Usuários.max_row+1):
                        valor_celula = Usuários.cell(row=usu, column=1).value
                        if valor_celula == cpf:
                            conftexto = str(Usuários.cell(row=usu, column=3).value) + '\n'
                            listaperfis.insert(END, conftexto)
                            if japegounome == 0:
                                nomeusu = str(Usuários.cell(row=usu, column=2).value)
                                japegounome = 1

                    lbl_nome = ctk.CTkLabel(master=frame_siscon, text='Usuário selecionado:', font=('Century Gothic bold', 22, 'bold'))
                    lbl_nome.place(x=170, y=110)
                    lbl_nome1 = ctk.CTkLabel(master=frame_siscon, text=nomeusu, font=('Century Gothic bold', 20))
                    lbl_nome1.place(x=170, y=140)

                    lbl_cpf = ctk.CTkLabel(master=frame_siscon, text='CPF do Usuário:', font=('Century Gothic bold', 22, 'bold'))
                    lbl_cpf.place(x=170, y=200)
                    lbl_cpf1 = ctk.CTkLabel(master=frame_siscon, text=cpf, font=('Century Gothic bold', 20))
                    lbl_cpf1.place(x=170, y=230)

                    lbl_text = ctk.CTkLabel(master=frame_siscon, text= f'Ao lado está a lista dos Perfis\nde acesso deste Usuário.', font=('Century Gothic bold', 18))
                    lbl_text.place(x=170, y=290)
                                
            cpf_value2 = StringVar()
            cpfentry = ctk.CTkEntry(master=frame6, textvariable=cpf_value2, width=200)
            cpfentry.place(x=177.5, y= 190)
            btn_consultar=ctk.CTkButton(master=frame6, text='Consultar'.upper(),command=consultar_usu, fg_color='#245983', hover_color='#063970').place(x=206.5, y=230)
                        

        frames_inicio() 
        criar_listas()

AppSOD()
